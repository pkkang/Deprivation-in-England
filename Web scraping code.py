#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Aug 16 10:57:39 2018

@author: pamillakang
"""
#GOES THROUGH A LIST OF URLS AND WEB SCRAPES THEM
#CREATES ONE SPREADSHEET OF THE WEB SCRAPED INFORMATION AND ONE SPREADSHEET WITH THE URLS THAT COULD NOT BE WEB SCRAPED

import openpyxl
from openpyxl import Workbook
from bs4 import BeautifulSoup
import requests

#define variables
#list of urls
urls=[]
#urls that have to be done manually
idek=[]
#the final list of all wards
wardarray=[]
#eventually a 2D array of the partys 
partyarray=[]
#temporary lists
wards1=[]
partys1=[]
wards1b=[]
partys1b=[]
wards1c=[]
partys1c=[]
#urls corresponding to each councillor
partyurl=[]
#urls corresponding to each ward
wardurl=[]
#overall representation to be calculated
overall=[]
#local authority district names
lad=[]
#open a spreadsheet with the urls and lads in 
wb2 = openpyxl.load_workbook('url and lad.xlsx')
ws2=wb2.active

#iterate over all the links and add them to a list
for i in range(1,ws2.max_row+1):
    urls.append(str(ws2.cell(row=i, column = 1).value))
    lad.append(str(ws2.cell(row=i, column=2).value))
#variable to check whether the url opened in python
opened =0;
#loops over the list of urls
for i in range(0,len(urls)):
    try:
        r = requests.get(urls[i])
        #r.content gets the html
        soup = BeautifulSoup(r.content, 'html.parser')
        opened =1;
    #if it doesn't open then print a message and put it into a list
    except:
        print('cant open url using r.requests ' + urls[i])
        idek.append(urls[i])
        continue
    
    if opened: # opened=1
        #print the url that it is scraping just so we know the code's progress
        print(urls[i])
        #set a variable that will later check whether the word parish is in any of the website text, since we don't want the lines about parishes
        parish=0;
        #try searching for mgThumbslist in the html
        classes = soup.find_all('div', attrs={'class': 'mgThumbsList'})
        #but if it doesn't find mgThumbsList then the list will be empty and let's search for something else in html
        if len(classes)==0:
            classes2 = soup.find_all('tr', attrs={'class': ['mgTableEvenRow', 'mgTableOddRow']})
            #if the website isn't in this format either, let's just do it manually
            if len(classes2)==0:
                idek.append(urls[i])
            #if the tag is in the html but in a different format then we shouldn't code that one either
            if len(classes2)==1:
                idek.append(urls[i])
            else: #for all the ones in a table format
                #loop over all the tr classes
                for j in range(0,len(classes2)):
                    if classes2[j].find('td') is not None:
                    #wards are the in the fourth td tag
                       wardsb=str(classes2[j].select_one('td:nth-of-type(4)')).lower()
                       # get rid of all the unwanted text from html
                       wardsb=wardsb.replace('<td>','')
                       wardsb=wardsb.replace('</td>','')
                       wardsb=wardsb.replace('&amp;','&')
                       wards1b.append(wardsb)
                       #councillor parties are in the third td
                       partysb=str(classes2[j].select_one('td:nth-of-type(3)'))
                       partysb=partysb.replace('<td>','')
                       partysb=partysb.replace('</td>','')
                       partysb=partysb.replace('<br/>',' ')
                       partysb=partysb.replace('&amp;','&')
                        #if the website does not give the councillor's party or the position is vacant, then we need to put not specified and should check these manually
                       if partysb is '':
                            partys1b.append('Not specified')
                       else: #if the councillors party is given then put it into a temporary list
                            partys1b.append(partysb)
                        #then put it into the total list of parties so that the temporary list can be emptied for the next website
                       for k in range(0,len(partys1b)):
                           partys1.append(partys1b[k])
                        #empty the list for the next website
                       partys1b=[]
                       #add the urls corresponding to each councillor to a list so that the councillors can be identified
                       partyurl.append(urls[i])  
                 #create a temporary array to get just a list of wards that are not repeated, unlike wards1b which is the wards corresponding to all the councillors
                ward=[]
                for x in wards1b:
                    if x not in ward:
                        ward.append(x)
                #add these to the final list of wards and also keep the url that corresponds to each ward so they can be identified
                for l in range(0,len(ward)):
                    wardarray.append(ward[l])
                    wardurl.append(urls[i])
                #we still need to keep a list of the wards that each councillor is in
                for l in range(0,len(wards1b)):
                    wards1.append(wards1b[l])
                #empty the array for the next website
                wards1b=[]

        if len(classes)==1: #for websites with one mgThumbsList tag
            for j in range (0,len(classes)):
                if classes[j].find('p') is not None:
                    #councillor information is in the li tags
                    eachli=classes[j].find_all('li')
                    for x in range(0,len(eachli)):
                        wardsc=str(eachli[x].select_one('p:nth-of-type(1)')).lower()
                        wardsc=wardsc.replace('<p>','')
                        wardsc=wardsc.replace('</p>','')
                        wardsc=wardsc.replace('&amp;','&')
                        wards1c.append(wardsc)
                         # but we dont want the text that is about the parishes
                         # parish info is usually in the second paragraph tag and parties are in the third p tag but this is inconsistent throughout the websites
                         # so we need the condition that there are three p tags and if the word parish is in the second tag, the parties are therefore the third p
                         # or if the website gives the parish in the second line but doesn't explicity say 'parish' then we set the variable parish to be true if each li tag has three p tags
                        if 'parish' in str(eachli[x].select_one('p:nth-of-type(2)')).lower() and eachli[x].select_one('p:nth-of-type(3)') is not None or eachli[x].select_one('p:nth-of-type(3)')is not None and parish==1:
                            partys=str(eachli[x].select_one('p:nth-of-type(3)'))
                            parish=1;
                        else: #for websites that dont give parishes
                            partys=str(eachli[x].select_one('p:nth-of-type(2)'))
                        partys=partys.replace('<p>','')
                        partys=partys.replace('</p>','')
                        partys=partys.replace('&amp;','&')

                        # not specified people
                        if partys is '':
                            partys1c.append('Not specified')
                        else:
                            partys1c.append(partys)
                        partyurl.append(urls[i])
                        for k in range(0,len(partys1c)):
                            partys1.append(partys1c[k])
                        partys1c=[]                                   
            ward=[]
            for x in wards1c:
                if x not in ward:
                    ward.append(x)     
            for l in range(0,len(ward)):
                    wardarray.append(ward[l])
                    wardurl.append(urls[i])
            for l in range(0,len(wards1c)):
                    wards1.append(wards1c[l])
            wards1c=[]
        else: #for websites with several mgThumbsList tags
            for j in range (0,len(classes)):
                if classes[j].find('p') is not None:
                    eachli=classes[j].find_all('li')
                    ward = str(classes[j].find('p')).lower()
                    wardurl.append(urls[i])
                    ward=ward.replace('<p>','')
                    ward=ward.replace('</p>','')
                    ward=ward.replace('&amp;','&')
                    wardarray.append(ward)
                    for x in range(0,len(eachli)):
                        wards=str(eachli[x].select_one('p:nth-of-type(1)')).lower()
                        wards=wards.replace('<p>','')
                        wards=wards.replace('</p>','')
                        wards=wards.replace('&amp;','&')
                        wards1.append(wards)
                        if 'parish' in str(eachli[x].select_one('p:nth-of-type(2)')).lower() and eachli[x].select_one('p:nth-of-type(3)') is not None or eachli[x].select_one('p:nth-of-type(3)')is not None and parish==1:
                            partys=str(eachli[x].select_one('p:nth-of-type(3)'))
                            parish=1
                        else:
                            partys=str(eachli[x].select_one('p:nth-of-type(2)'))
                        partys=partys.replace('<p>','')
                        partys=partys.replace('</p>','')
                        partys=partys.replace('&amp;','&')
                        if partys is '':
                            partys1.append('Not specified')
                        else:
                            partys1.append(partys)
                        partyurl.append(urls[i])

#make a new workbook to put the scraped data in
wb=Workbook()
ws=wb.active
#make another workbook to put the errors in
wb3=Workbook()
ws3=wb3.active

#make a 2D party array list
#the first index of partyarray corresponds to the ALL the councillors in a certain ward
#the second index of partyarray corresponds to a single councillor in that ward
#eg wardarray[1] = 'Withington' ... this is an example ward
#partyarray[1]=['Labour','Conservative','Lib Dem'] ... these are all the councillors within the withington ward
#partyarray[1][1]='Conservative' ... this is just one of the councillors in the withington ward

for m in range(0,len(wardarray)):
    ws.cell(row = m+2, column=1).value=wardarray[m] #write the wards to the spreadsheet
    partyarray.append([])
    for n in range (0,len(wards1)):
        if wardarray[m]==wards1[n] and wardurl[m]==partyurl[n]:
            partyarray[m].append(partys1[n])

#calculate the overall representation of each ward by counting the number of times each party appears
for x in range(0,len(wardarray)):
    for k in range (0,len(partyarray[x])):
        ws.cell(row=x+2, column =k+2).value = partyarray[x][k]
    if sum('conservative'in s.lower() for s in partyarray[x])>=2:
        overall.append('Conservative')
    elif len(partyarray[x])==1 and sum('conservative'in s.lower() for s in partyarray[x])==1:
        overall.append('Conservative')
    elif sum('labour'in s.lower() for s in partyarray[x])>=2:
        overall.append('Labour')
    elif len(partyarray[x])==1 and sum('labour'in s.lower() for s in partyarray[x])==1:
         overall.append('Labour')
    elif sum('liberal democrat'in s.lower() for s in partyarray[x])>=2:
        overall.append('Liberal Democrat')
    elif len(partyarray[x])==1 and sum('liberal democrat'in s.lower() for s in partyarray[x])==1:
         overall.append('Liberal Democrat')
    elif sum('green'in s.lower() for s in partyarray[x])>=2:
        overall.append('Green')
    elif len(partyarray[x])==1 and sum('green'in s.lower() for s in partyarray[x])==1:
        overall.append('Green')
    elif sum('uk'in s.lower() for s in partyarray[x])>=2:
        overall.append('UKIP')
    elif len(partyarray[x])==1 and sum('uk'in s.lower() for s in partyarray[x])==1:
        overall.append('UKIP')
    elif len(partyarray[x])==2 and partyarray[x][0]!=partyarray[x][1]:
        overall.append('NOC')
    elif len(partyarray[x])==3 and partyarray[x][0]!=partyarray[x][1]and partyarray[x][0]!=partyarray[x][2]and partyarray[x][2]!=partyarray[x][1]:
        overall.append('NOC')
    else:
        overall.append('Other')

#write the information to the spreadsheet
for j in range(0,len(wardurl)):
    ws.cell(row=j+2,column =7).value=wardurl[j]
    for m in range(0,len(urls)):
        if wardurl[j]==urls[m]:
            ws.cell(row=j+2, column=6).value=lad[m]
for j in range(0,len(overall)):
    ws.cell(row=j+2,column =5).value=overall[j]

#write the manual websites to a different spreadsheet
#i'm just going to put a reminder at the bottom about things that have to be done in the web scraped data 
idek.append('Dont forget to check the not specified people in the web scraped data as it either means there is a vacant spot or a councillor doesnt give their party - City of London will have lots of Not Specified columns')
idek.append('Check for the word elected in the web scraped data as there is one website that includes the date that the councillors are elected next to their political party')
for j in range(0,len(idek)):
    ws3.cell(row=j+2,column =1).value=idek[j]

#titles in the spreadsheets
ws.cell(row=1, column=1).value='Ward Name'
ws.cell(row=1, column=2).value='COUNC 1'
ws.cell(row=1, column=3).value='COUNC 2'
ws.cell(row=1, column=4).value='COUNC 3'
ws.cell(row=1, column=7).value='URL'
ws.cell(row=1, column=6).value='LAD'
ws.cell(row=1, column=5).value='REPR'
ws3.cell(row=1, column=1).value='Have to do Manually'

wb.save('Web scraped councillors.xlsx')
wb3.save('Errors and manual labour.xlsx')
