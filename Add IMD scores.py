#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Aug 18 22:21:12 2018

@author: pamillakang
"""
#ADDS THE IMD15 SCORES ONTO THE NEW COUNCILLOR DATABASE

import openpyxl
from openpyxl import Workbook
#i am using the old database to get the IMD15 for each LSOA, but the IMD15 information can be found separately 
#THE DATABASES NEED TO BE ORDERED BY ALPHABETICAL LSOA CODE FOR THIS TO WORK
#when both the new and old data base are ordered alphabetically by LSOA code, then the deprivation score can just be written onto the new database line by line
#rather than searching both databases to match the LSOA codes and then matching the IMD score, which would take ages
wb = openpyxl.load_workbook('DEPVSREP240418.xlsx')
ws=wb.active
wb2= openpyxl.load_workbook('Total data.xlsx')
ws2=wb2.active
wb3=Workbook()
ws3=wb3.active

for i in range(1,ws2.max_row+1):
    ws3.cell(column=1,row=i).value=ws2.cell(column=1,row=i).value
    ws3.cell(column=2,row=i).value=ws2.cell(column=2,row=i).value
    ws3.cell(column=3,row=i).value=ws2.cell(column=3,row=i).value
    ws3.cell(column=4,row=i).value=ws2.cell(column=4,row=i).value
    ws3.cell(column=5,row=i).value=ws2.cell(column=5,row=i).value
    ws3.cell(column=6,row=i).value=ws2.cell(column=6,row=i).value
    ws3.cell(column=7,row=i).value=ws2.cell(column=7,row=i).value
    ws3.cell(column=8,row=i).value=ws2.cell(column=8,row=i).value
    ws3.cell(column=9,row=i).value=ws.cell(column=8,row=i).value
    ws3.cell(column=10,row=i).value=ws2.cell(column=10,row=i).value
    ws3.cell(column=11,row=i).value=ws2.cell(column=11,row=i).value
    ws3.cell(column=12,row=i).value=ws2.cell(column=12,row=i).value
    ws3.cell(column=13,row=i).value=ws2.cell(column=13,row=i).value
    ws3.cell(column=14,row=i).value=ws2.cell(column=14,row=i).value

wb3.save('Deprivation in England final.xlsx')