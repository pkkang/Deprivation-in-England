#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Aug 16 13:13:39 2018

@author: pamillakang
"""
#CREATES A DATABASE OF COUNCILLORS IN ENGLAND USING THE ONS WARD NAMES AND THE WEB SCRAPED DATA
import openpyxl
from openpyxl import Workbook

# open the workbook from the ONS with the LSOA codes and wards
wb = openpyxl.load_workbook('Lower_Layer_Super_Output_Area_2011_to_Ward_to_LAD_May_2018_Lookup_in_England_and_Wales.xlsx')
ws=wb.active
# open the final list of web scraped councillors and manually done websites - all the councillor data
wb2= openpyxl.load_workbook('Final councillors list.xlsx')
ws2=wb2.active
wb3=Workbook()
ws3=wb3.active

wards=[]
wards2=[]
#put the ONS wards into a list
for x in range(2,ws.max_row+1):
    wards.append(str(ws.cell(row=x,column=4).value).lower())
#put the web scraped wards into a list
for y in range(2,ws2.max_row+1):
    wards2.append(str(ws2.cell(row=y, column=1).value).lower())

#remove all the unimportant punctuation from the wards 
#so that the ONS wards and web scraped wards are more consistent and can therefore be matched
for k in range(0,len(wards2)):
    wards2[k]=wards2[k].replace('&','')
    wards2[k]=wards2[k].replace(' and ','')
    wards2[k]=wards2[k].replace("'",'')
    wards2[k]=wards2[k].replace("’",'')
    wards2[k]=wards2[k].replace("-",'')
    wards2[k]=wards2[k].replace(".",'')
    wards2[k]=wards2[k].replace(" ",'')
    wards2[k]=wards2[k].replace(",",'')
    wards2[k]=wards2[k].replace("(",'')
    wards2[k]=wards2[k].replace(")",'')
    wards2[k]=wards2[k].replace("/",'')
for i in range(0,len(wards)):
    wards[i]=wards[i].replace('&','')
    wards[i]=wards[i].replace(' and ','')
    wards[i]=wards[i].replace('-','')
    wards[i]=wards[i].replace("'",'')
    wards[i]=wards[i].replace("’",'')
    wards[i]=wards[i].replace(".",'')
    wards[i]=wards[i].replace(" ",'')
    wards[i]=wards[i].replace(",",'')
    wards[i]=wards[i].replace("(",'')
    wards[i]=wards[i].replace(")",'')
    wards[i]=wards[i].replace("/",'')
    
#match the wards that are in the same local authority district and therefore write their info to a spreadsheet
    for k in range(0,len(wards2)):
        if wards[i] in wards2[k] and str(ws.cell(row=i+2,column=7).value)==str(ws2.cell(row=k+2,column=6).value) or wards2[k] in wards[i] and str(ws.cell(row=i+2,column=7).value)==str(ws2.cell(row=k+2,column=6).value):
            ws3.cell(column=1,row=i+2).value=ws.cell(column=1,row=i+2).value
            ws3.cell(column=2,row=i+2).value=ws.cell(column=2,row=i+2).value
            ws3.cell(column=3,row=i+2).value=ws.cell(column=3,row=i+2).value
            ws3.cell(column=4,row=i+2).value=ws.cell(column=4,row=i+2).value
            ws3.cell(column=5,row=i+2).value=ws.cell(column=5,row=i+2).value
            ws3.cell(column=6,row=i+2).value=ws.cell(column=6,row=i+2).value
            ws3.cell(column=7,row=i+2).value=ws.cell(column=7,row=i+2).value
            ws3.cell(column=8,row=i+2).value=ws.cell(column=8,row=i+2).value
            ws3.cell(column=10,row=i+2).value=ws2.cell(column=2,row=k+2).value
            ws3.cell(column=11,row=i+2).value=ws2.cell(column=3,row=k+2).value
            ws3.cell(column=12,row=i+2).value=ws2.cell(column=4,row=k+2).value
            ws3.cell(column=13,row=i+2).value=ws2.cell(column=5,row=k+2).value
            ws3.cell(column=14,row=i+2).value=ws2.cell(column=7,row=k+2).value
#write the titles
ws3.cell(row=1, column=1).value='LSOA11CD'
ws3.cell(row=1, column=2).value='LSOA11NM'
ws3.cell(row=1, column=3).value='WD18CD'
ws3.cell(row=1, column=4).value='WD18NM'
ws3.cell(row=1, column=5).value='WD18NMW'
ws3.cell(row=1, column=6).value='LAD18CD'
ws3.cell(row=1, column=7).value='LAD18NM'
ws3.cell(row=1, column=8).value='FID'
ws3.cell(row=1, column=9).value='IMD15'
ws3.cell(row=1, column=10).value='COUNC 1'
ws3.cell(row=1, column=11).value='COUNC 2'
ws3.cell(row=1, column=12).value='COUNC 3'
ws3.cell(row=1, column=13).value='REPR'
ws3.cell(row=1, column=14).value='URL'

#if a ward hasn't been matched because they are spelt differently on their website to the ONS, we need to do these manually
print('These wards need to be checked manually because they are spelt differently on their website to the ONS data')
for x in range(1,ws3.max_row+1): 
    if ws3.cell(row=x, column=10).value is None and 'E' in str(ws.cell(column=1,row=x).value):
        print(ws.cell(column=4,row=x).value +'   ' +ws.cell(row=x,column=7).value)
        
#we still need to write the wards that didn't manage to match up to the database
#this includes the welsh wards but these can be deleted after if we want
for x in range(1, ws.max_row+1):
    if ws3.cell(row=x,column=1).value is None:
        ws3.cell(row=x,column=1).value=ws.cell(row=x, column=1).value
        ws3.cell(row=x,column=2).value=ws.cell(row=x, column=2).value
        ws3.cell(row=x,column=3).value=ws.cell(row=x, column=3).value
        ws3.cell(row=x,column=4).value=ws.cell(row=x, column=4).value
        ws3.cell(row=x,column=5).value=ws.cell(row=x, column=5).value
        ws3.cell(row=x,column=6).value=ws.cell(row=x, column=6).value
        ws3.cell(row=x,column=7).value=ws.cell(row=x, column=7).value
        ws3.cell(row=x,column=8).value=ws.cell(row=x, column=8).value

wb3.save('Total data wed.xlsx')

        
        
        
        
        
